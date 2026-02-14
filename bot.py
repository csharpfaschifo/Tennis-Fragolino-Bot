# ============================================================================
# TELEGRAM BOT - TENNIS MATCH PROCESSOR
# Bot che riceve foto di match e aggiorna automaticamente l'Excel su Google Drive
# ============================================================================

import os
import re
import unicodedata
import pandas as pd
import numpy as np
from PIL import Image, ImageEnhance, ImageFilter
from openpyxl import load_workbook
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
from googleapiclient.discovery import build
from google.oauth2 import service_account
from googleapiclient.http import MediaFileUpload
import json
import traceback
from difflib import SequenceMatcher
from googleapiclient.http import MediaIoBaseDownload
import io
import shutil
import pytesseract
pytesseract.pytesseract.tesseract_cmd = "/usr/bin/tesseract"
import subprocess
print("DEBUG: which tesseract =", shutil.which("tesseract"))
print("DEBUG: tesseract version:")
print(subprocess.run(["/usr/bin/tesseract", "--version"], capture_output=True, text=True).stdout)

# tesseract_path = shutil.which("tesseract")
# if tesseract_path:
#     pytesseract.pytesseract.tesseract_cmd = tesseract_path
# else:
#     raise RuntimeError("❌ Tesseract non trovato nel PATH")


# ============================================================================
# CONFIGURAZIONE
# ============================================================================

# Il tuo token del bot (lo hai ottenuto da BotFather)
TELEGRAM_TOKEN = os.environ["TELEGRAM_TOKEN"]
ALLOWED_USERS = {
    8512235231,
    49697387
}


# Path del file Excel su Google Drive
EXCEL_FILE_ID = "ID_FILE_GOOGLE_DRIVE"  # ← Lo otterrai dopo
EXCEL_LOCAL_PATH = "Database_Tennis.xlsx"
SCOPES = ["https://www.googleapis.com/auth/drive"]
creds = service_account.Credentials.from_service_account_info(
    json.loads(os.environ["GOOGLE_APPLICATION_CREDENTIALS"]),
    scopes=SCOPES
)

drive_service = build("drive", "v3", credentials=creds)
EXCEL_FILE_ID = "1WaFe87w0bR2WlweXI-pUFupwUfl8g9g5"

# ============================================================================
# LISTA TENNISTI
# ============================================================================

tennisti = [
    "Carlos Alcaraz", "Jannik Sinner", "Alexander Zverev", "Novak Djoković", "Lorenzo Musetti", "Taylor Fritz", "Alex de Minaur", "Félix Auger-Aliassime", "Ben Shelton", "Alexander Bublik",
    "Daniil Medvedev", "Jack Draper", "Casper Ruud", "Alejandro Davidovich Fokina", "Andrey Rublev", "Jakub Menšík", "Holger Rune", "Karen Khachanov", "Francisco Cerúndolo", "Flavio Cobolli",
    "Jiří Lehečka", "Luciano Darderi", "Learner Tien", "Tommy Paul", "Valentin Vacherot", "Cameron Norrie", "Tallon Griekspoor", "Arthur Rinderknech", "Brandon Nakashima", "Frances Tiafoe",
    "Tomáš Macháč", "Stefanos Tsitsipas", "João Fonseca", "Sebastián Báez", "Corentin Moutet", "Ugo Humbert", "Jaume Munar", "Gabriel Diallo", "Zizou Bergs", "Denis Shapovalov",
    "Alex Michelsen", "Arthur Fils", "Grigor Dimitrov", "Nuno Borges", "Jenson Brooksby", "Fábián Marozsán", "Alexandre Müller", "Márton Fucsovics", "Alexei Popyrin", "Daniel Altmaier",
    "Sebastian Korda", "Tomás Martín Etcheverry", "Camilo Ugo Carabelli", "Kamil Majchrzak", "Adrian Mannarino", "Giovanni Mpetshi Perricard", "Valentin Royer", "Matteo Berrettini", "Marcos Giron", "Lorenzo Sonego",
    "Marin Čilić", "Damir Džumhur", "Francisco Comesaña", "Térence Atmane", "Botic van de Zandschulp", "Miomir Kecmanović", "Reilly Opelka", "Eliot Spizzirri", "Raphaël Collignon", "Hubert Hurkacz",
    "Alejandro Tabilo", "Matteo Arnaldi", "Mariano Navone", "Ethan Quinn", "Arthur Cazaux", "Filip Misolic", "Quentin Halys", "Aleksandar Kovačević", "Hamad Medjedović", "Emilio Nava",
    "Jan Lennard Struff", "Juan Manuel Cerúndolo", "James Duckworth", "Alexander Shevchenko", "Roberto Bautista Agut", "Jesper de Jong", "Jacob Fearnley", "Aleksandar Vukic", "Cristian Garín", "Yannick Hanfmann",
    "Laslo Djere", "Thiago Agustín Tirante", "Adam Walton", "Pedro Martínez", "Vít Kopřiva", "Ignacio Buse", "Luca Nardi", "Dalibor Svrčina", "Hugo Gaston", "Carlos Taberner",
    "Adolfo Daniel Vallejo", "Alexander Blockx", "Pablo Carreño Busta", "Stan Wawrinka", "Mattia Bellucci", "Patrick Kypson", "Zachary Svajda", "Jordan Thompson", "Tristan Schoolkate", "Mackenzie McDonald",
    "Rinky Hijikata", "Tomás Barrios Vera", "Shintaro Mochizuki", "Francesco Maestrelli", "Román Andrés Burruchaga", "Chun Hsin Tseng", "David Goffin", "Otto Virtanen", "Christopher O'Connell", "Elmer Møller",
    "Dino Prižmić", "Dušan Lajović", "Rafael Jódar", "Jan Choinski", "Luca Van Assche", "Vilius Gaubas", "Kyrian Jacquet", "Borna Ćorić", "Martin Damm", "Benjamin Bonzi",  
    "Nicolai Budkov Kjaer", "Billy Harris", "Marco Trungelliti", "Sebastian Ofner", "Coleman Wong", "Nikoloz Basilashvili", "Titouan Droguet", "Chris Rodesch", "Brandon Holt", "Yunchaokete Bu",
    "Nicolás Jarry", "Sho Shimabukuro", "Andrea Pellegrino", "Moez Echargui", "Yibing Wu", "Jaime Faria", "Liam Draxl", "Lukáš Klein", "Michael Zheng", "Dane Sweeny",
    "Arthur Fery", "Giulio Zeppieri", "Martin Landaluce", "Arthur Géa", "Yoshihito Nishioka", "Colton Smith", "Mark Lajal", "Francesco Passaro", "Ugo Blanchet", "Matteo Gigante",
    "Gaël Monfils", "Stefano Travaglia", "Luka Mikrut", "Daniel Mérida", "Henrique Rocha", "Yosuke Watanuki", "Hugo Dellien", "Harold Mayot", "Guy Den Ouden", "Daniil Glinka",
    "Pierre Hugues Herbert", "Jurij Rodionov", "Leandro Riedi", "Rei Sakamoto", "Roberto Carballés Baena", "Zsombor Piros", "Nicolás Mejía", "Alex Bolt", "Vitaliy Sachko", "Alex Barrena",
    "Jack Pinnington Jones", "Juan Pablo Ficovich", "Jay Clarke", "Roman Safiullin", "Elias Ymer", "Lloyd Harris", "Daniel Elahi Galán", "Hugo Grenier", "Zdeněk Kolář", "Daniel Evans",
    "Timofey Skatov", "Bernard Tomic", "Nishesh Basavareddy", "Justin Engel", "Jérôme Kym", "Federico Agustin Gomez", "Thiago Seyboth Wild", "Clément Chidekh", "August Holmgren", "Alex Molčan",
    "Dan Added", "Álvaro Guillén Meza", "Jason Kubler", "George Loffhagen", "Pablo Llamas Ruiz", "Lorenzo Giustino", "João Lucas Reis Da Silva", "James McCabe", "Thiago Monteiro", "Gilles Arnaud Bailly",
    "Juan Carlos Prado Angelo", "Remy Bertola", "Clément Tabur", "Ilia Simakin", "Oliver Crawford", "Tristan Boyer", "Marco Cecchinato", "Gonzalo Bueno", "Aidan Mayo", "Alec Deckers",
    "Mees Rottgering", "Gerard Campana Lee", "Yassine Dlimi", "Shintaro Imai", "Damien Wenger", "Lorenzo Sciahbasi", "Maxence Beaugé", "Louis Weßels", "Aziz Ouakaa", "Máté Valkusz",
    "Jacopo Vasami", "Karol Filar", "Kuan Yi Lee", "S D Prajwal Dev", "Mario Gonzalez Fernandez", "Imanol López Morillo", "Nicolas Moreno De Alboran", "Cem İlkel", "Kenta Miyoshi", "Yuta Kawahashi",
    "Juan Cruz Martin Manzano", "Denis Klok", "Daniel Siniakov", "Jakub Nicod", "Maxime Chazal", "Dennis Novak", "Ye Cong Mo", "Felix Balshaw", "William Grant", "Andrew Paulson",
    "Christian Sigsgaard", "Yusuke Kusuhara", "Rigele Te", "Santiago De La Fuente", "Massimo Giunta", "Johan Alexander Rodriguez", "Naoki Nakagawa", "Francesco Forti", "Felix Corwin", "Facundo Juárez",
    "Semen Pankin", "Ryan Fishback", "Kyle Kang", "Aidan Kim", "Alafia Ayeni", "Toby Kodat", "Igor Kudriashov", "Stijn Slump", "Jan Kumstát", "Federico Iannaccone",
    "Giovanni Oradini", "Duncan Chan", "Karue Sell", "James Watt", "Cruz Hewitt", "Radu David Turcanu", "Giuseppe La Vela", "Luca Castagnola", "Yurii Dzhavakian", "Michael Agwi",
    "Tiago Torres", "Timofei Derepasko", "Pierre Delage", "Chase Ferguson", "Sergi Perez Contri", "Dimitris Sakellaridis", "Alexander Donski", "Enzo Couacaud", "Pierluigi Basile", "Alex Knaff",
    "Alejandro Manzanera Pertusa", "Abel Forger", "Ryan Dickerson", "Marcel Zieliński", "Tim Handel", "Luca Staeheli", "Matisse Bobichon", "Rodrigo Alujas", "Michael Bassem Sobhy", "Jozef Kovalík",
    "Anton Matusevich", "Arthur Reymond", "Yuta Kikuchi", "Mitsuki Wei Kang Leong", "Cooper Williams", "Lucas Gerch", "Daniel Antonio Núñez", "Yaroslav Demin", "Axel Garcian", "Daniil Ostapenkov",
    "Millen Hurrion", "Tianhui Zhang", "Kenny de Schepper", "Leo Vithoontien", "Kyle Edmund", "Constant Lestienne", "Pablo Martinez Gomez", "Tuncay Duran", "Peter Bertran", "Kaylan Bigun",
    "Sora Fukuda", "Sidharth Rawat", "Gabriel Ghetu", "Ali Yazdani", "Strong Kirchheimer", "Pedro Cachín", "José Pereira", "Dinko Dinev", "Yanis Ghazouani Durand", "Diego Augusto Barreto Sánchez",
    "Aslan Karatsev", "Blu Baker", "Pedro Rodrigues", "Uisung Park", "Amaury Raynel", "Maik Steiner", "Mika Petković", "JiSung Nam", "John Echeverria", "Matyáš Černý",
    "Maxence Bertimon", "Matthew Summers", "Matteo Covato", "Ignacio Monzón", "Evan Zhu", "Conner Huertas del Pino", "Marlon Vankan", "Josip Šimundža", "Kai Wehnelt", "Nicolas Tepmahc",
    "Michiel De Krom", "Maximilian Marterer", "Nick Kyrgios", "Juan Carlos Aguilar", "Tomás Farjat", "Adrien Gobat", "Samuel Alejandro Linde Palacios", "Yanaki Milev", "Ben Jones", "Alexis Gautier",
    "César Bouchelaghem", "Léo Raquillet", "Lorenzo Rottoli", "Tomas Curras Abasolo", "Guillaume Dalmasso", "Samuel Heredia", "Pierre Yves Bailly", "Fryderyk Lechno-Wasiutyński", "Théo Papamalamis", "Alessandro Pecci",
    "Mwendwa Mbithi", "Keisuke Saitoh", "Mikalai Haliak", "Daniel De Jonge", "Matt Kuhar", "Attila Boros", "Kris Van Wyk", "Aleksandr Braynin", "Jack Anthrop", "Karim Bennani",
    "Victor Lilov", "Thanapet Chanta", "Wilson Leite", "Francisco Rocha", "John Sperle", "Menelaos Efstathiou", "Leonardo Rossi", "Lucas Andrade Da Silva", "Oscar Weightman", "Roger Pascual Ferra",
    "Nikita Bilozertsev", "Ivan Ivanov", "Matyas Fuele", "Kārlis Ozoliņš", "Emile Hudd", "Alec Beckley", "M Rifki Fitriadi", "Bruno Fernandez", "Benjamin Torrealba", "Colin Sinclair",
    "Toufik Sahtali", "Fausto Tabacco", "Iannis Miletich", "Oscar Jose Gutierrez", "Wishaya Trongcharoenchaikul", "Ilya Ivashka", "Evgeny Karlovskiy", "Gustavo Ribeiro De Almeida", "Maximo Zeitune", "Linang Xiao",
    "Bautista Vilicich", "Martin Van Der Meerschen", "Amir Omarkhanov", "Pedro Araújo", "Albert Pedrico Kravtsov", "Marko Miladinović", "Pedro Rodenas", "Oliver Bonding", "Duck-hee Lee", "Shinji Hazawa",
    "Alejandro Juan Mano", "Artur Kukasian", "Evgenii Tiurnev", "Luca Preda", "Ajeet Rai", "Alan Magadan", "Joshua Charlton", "Taiyo Yamanaka", "Egor Gerasimov", "Petros Tsitsipas",
    "Yannik Kelm", "Jorge Plans", "Maximilian Homberg", "Michel Hopp", "Izan Almazan Valiente", "Niccolo Catini", "Alexandre Aubriot", "Samir Hamza Reguig", "Martin Kližan", "Gian Luca Tanner"#,
    # "Matt Ponchet", "Pavel Lagutin", "Kristijan Juhas", "Alejandro Turriziani Álvarez", "Jeffrey Von Der Schulenburg", "Mikael Ymer", "Dong Ju Kim", "Arklon Huertas Del Pino Cordova", "Grégoire Barrère", "Julio César Porras",
    # "Rafael Izquierdo Luque", "Nikola Slavic", "Vincent Marysko", "Thiago Cigarran", "Jacob Brumm", "Patrick Schoen", "Hoyoung Roh", "Marc Van Der Merwe", "Scott Jones", "Xiaofei Wang",
    # "Ignacio Parisca Romera", "John Hallquist Lithen", "Amit Vales", "Taisei Ichikawa", "Luca Pow", "Luca Wiedenmann", "Peter Makk", "Andreja Petrovic", "Giorgio Tabacco", "Charlie Robertson",
    # "Ivan Nedelko", "Jumpei Yamasaki", "Lautaro Agustin Falabella", "Jakub Filip", "Mert Naci Türker", "Noah Schachter", "Manuel Plunger", "Emon Van Loben Sels", "Kalin Ivanovski", "Lorenzo Bocchi",
    # "Joao Victor Couto Loureiro", "Stefano D'Agostino", "Niklas Schell", "Seydina André", "Andrea Fiorentini", "Dragos Nicolae Cazacu", "Diego Schwartzman", "Thanasi Kokkinakis", "Nicolas Ian Kotzen", "Aoran Wang",
    # "Tiago Cação", "Micah Braswell", "Jack Kennedy", "Manish Sureshkumar", "Aleksandr Lobanov", "Daisuke Sumizawa", "Anas Mazdrashki", "Marko Maksimović", "Trey Hilderbrand", "Ilya Snițari",
    # "Samuel Vincent Ruggeri", "Shunsuke Nakagawa", "Tomas Serrano Luis", "Marcus Walters", "Nitin Kumar Sinha", "Alexander Klintcharov", "Finn Murgett", "Romain Faucon", "Ezekiel Clark", "Nicolas Jadoun",
    # "Matthew Forbes", "Theodore Dean", "Calvin Müller", "Anton Shepp", "Isaiah Strode", "Enzo Kohlmann De Freitas", "Lucas Bouquet", "Kasra Rahmani", "Marc Majdandzic", "Benjamin Pietri",
    # "Carles Hernández", "Nikolai Barsukov", "Nikita Mashtakov", "Kosuke Ogura", "Orlando Luz", "Paul Inchauspe", "Jacob Bradshaw", "Salvador Price", "Ryotaro Taguchi", "Yaojie Zeng",
    # "Fernando Cavallo", "Miles Jones", "Dev Javia", "Naoya Honda", "Evan Bynoe", "Jesse Delaney", "Zoran Ludoški", "Ryan Colby", "Rudy Quan", "Pablo Masjuan Ginel",
    # "Stijn Paardekooper", "Adan Freire Da Silva", "Nicolas Zanellato", "Lorenzo Angelini", "Mihai Razvan Marinescu", "Grigoriy Lomakin", "Steven Diez", "Spencer Johnson", "Yun seong Chung", "Juan Sebastian Osorio",
    # "Qian Sun", "Brian Bozemoj", "Younes Lalami", "Kristjan Tamm", "Pietro Marino", "Diego Fernandez Flores", "Michael Zhu", "David Pichler", "Johannes Ingildsen", "Ozan Baris",
    # "Taym Al Azmeh", "Oskar Brostrom Poulsen", "Jea Moon Lee", "Pietro Romeo Scomparin", "Thantub Suksumrarn", "Ewen Lumsden", "Justin Schlageter", "Gabriele Bosio", "Maxim Shin", "Ezequiel Monferrer",
    # "Jeremy Gschwendtner", "Axel Nefve", "Nicolás Barrientos", "Dante Pagani", "Ryo Tabata", "Benjamin Winter Lopez", "Joris De Loore", "Markus Malaszszak", "Karl Friberg", "Bekkhan Atlangeriev",
    # "Marcello Serafini", "Tomás Martínez", "Benjamin Thomas George", "Daniel Salazar", "Lorenzo Lorusso", "Finn Bass", "Nicolás Villalón", "Žiga Šeško", "Andrew Delgado", "Kuang Qing Xu",
    # "Aleksa Ćirić", "Mateo Barreiros Reyes", "Suk Hyun Choo", "Dominique Rolland", "Mathias Bourgue", "Gavin Young", "Kiranpal Pannu", "Alejo Lorenzo Lingua Lavallén", "Leonardo Aboian", "William Rejchtman Vinciguerra",
    # "Zach Stephens", "Enzo Wallart", "Peter Buldorini", "Tymur Bieldiugin", "Jesse Flores", "Mikael Arseneault", "Chirag Duhan"
]

player_surname = [x.split()[-1].lower() for x in tennisti]
# Mappa: cognome_normalizzato → cognome_originale
def normalizza_nome(nome: str) -> str:
    if not nome:
        return ""

    # 1. lowercase
    nome = nome.lower()

    # 2. decomposizione unicode (accenti → lettere base)
    nome = unicodedata.normalize("NFKD", nome)
    nome = "".join(c for c in nome if not unicodedata.combining(c))

    # 3. rimuove tutto ciò che non è lettera
    nome = re.sub(r"[^a-z]", "", nome)

    return nome
    
mappa_cognomi = {
    normalizza_nome(cognome): cognome
    for cognome in player_surname
}

# ============================================================================
# FUNZIONI DI PREPROCESSING IMMAGINE
# ============================================================================

def gray_scale_img(img):
    """Prepara l'immagine per OCR"""
    img = img.convert("L")
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(2.0)
    sharpener = ImageEnhance.Sharpness(img)
    img = sharpener.enhance(2.0)
    img = img.filter(ImageFilter.MedianFilter(size=3))
    arr = np.array(img)
    threshold = arr.mean() + 20 # cambiato da 15 a 20
    arr = np.where(arr > threshold, 255, 0).astype("uint8")
    img = Image.fromarray(arr)
    return img

# ============================================================================
# FUNZIONI DI ESTRAZIONE DATI (dal tuo notebook)
# ============================================================================

def separa_maiuscole(testo: str) -> str:
    """
    Inserisce uno spazio prima di ogni lettera maiuscola
    quando è attaccata a una minuscola (HamadMededovic).
    """
    testo = re.sub(r'([a-zà-ž])([A-ZÀ-Ž])', r'\1 \2', testo)
    return testo

def similarita(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio()

def trova_cognome_nella_lista(lista_tennisti, candidati):
    trovati = []
    usati = set()

    for nome in candidati:
        nome_norm = normalizza_nome(nome)

        if len(nome_norm) < 3:
            continue

        miglior_match = None
        miglior_score = 0.0

        for cognome_norm, cognome_reale in mappa_cognomi.items():
            if cognome_reale in usati:
                continue

            # MATCH ESATTO
            if nome_norm == cognome_norm:
                miglior_match = cognome_reale
                miglior_score = 1.0
                break

            # MATCH PARZIALE
            if nome_norm in cognome_norm or cognome_norm in nome_norm:
                score = 0.9
            else:
                score = similarita(nome_norm, cognome_norm)

            if score > miglior_score:
                miglior_score = score
                miglior_match = cognome_reale

        # soglia di sicurezza
        if miglior_score >= 0.82 and miglior_match:
            trovati.append(miglior_match)
            usati.add(miglior_match)

        if len(trovati) == 2:
            break
    
    if len(trovati) == 1: # se ne trovi solo uno → avversario non riconosciuto
        trovati.append("NON_RICONOSCIUTO")

    return trovati

# def estrai_game_da_testo(testo, giocatori):
#     testo = testo.lower()
#     righe = testo.split("\n")

#     g1 = normalizza_nome(giocatori[0]) if giocatori[0] != "NON_RICONOSCIUTO" else None
#     g2 = normalizza_nome(giocatori[1]) if giocatori[1] != "NON_RICONOSCIUTO" else None

#     game_g1, game_g2 = [], []

#     for riga in righe:
#         riga_norm = normalizza_nome(riga)
#         numeri = list(map(int, re.findall(r'\d+', riga)))

#         # if g1 and g1 in riga_norm:
#         #     game_g1 = numeri
#         # elif g2 and g2 in riga_norm:
#         #     game_g2 = numeri
#         if not numeri:
#             continue

#         if g1 and similarita(riga_norm, g1) > 0.7:
#             game_g1 = numeri
#         elif g2 and similarita(riga_norm, g2) > 0.7:
#             game_g2 = numeri

#     return game_g1, game_g2

def estrai_game_da_testo(testo: str):
    """
    Estrae i game dai punteggi del match leggendo SOLO
    i numeri FINALI delle righe-score.
    """
    righe = [r.strip() for r in testo.split("\n") if r.strip()]
    risultati = []
    for riga in righe:
        riga_pulita = re.sub(r"[^\w\s]", " ", riga).lower()
        # scarta righe statistiche
        if any(k in riga_pulita for k in [
            "ace", "doppi", "falli", "%", "break", "tiebreak"
        ]):
            continue
        # match SOLO numeri alla fine
        match = re.search(r'(\d+(?:\s+\d+)*)\s*$', riga_pulita)
        if not match:
            continue

        numeri = list(map(int, match.group(1).split()))
        # devono essere almeno 2 e plausibili per tennis
        if len(numeri) < 2:
            continue
        if any(n > 7 for n in numeri):
            continue
        risultati.append(numeri)
        if len(risultati) == 2:
            break
    # fallback sicuro
    if len(risultati) == 1:
        risultati.append([])
    if len(risultati) == 0:
        return [], []
    return risultati[0], risultati[1]

def calcola_tie_break(game_g1, game_g2): 
    tie_breaks = 0 
    for g1, g2 in zip(game_g1, game_g2):
        if g1 + g2 >= 13: 
            tie_breaks += 1 
    return tie_breaks

def estrai_statistiche(testo):
    ace_match = re.search(r'(\d+)\s+Ace\s+(\d+)', testo, re.IGNORECASE)
    ace = [int(ace_match.group(1)), int(ace_match.group(2))] if ace_match else [0, 0]
    
    df_match = re.search(r'(\d+)\s+Doppi falli\s+(\d+)', testo, re.IGNORECASE)
    doppi_falli = [int(df_match.group(1)), int(df_match.group(2))] if df_match else [0, 0]
    # br1, br2 = "0/0", "0/0" # per evitare conflitti in caso di 1 giocatore mancante

    # for el in testo.lower().split("\n"):
    #     if "break point" in el:
    #         br1 = el.split("break point")[0].replace(" ", "")
    #         br2 = el.split("break point")[1].replace(" ", "")
    break_match = re.search(r'(\d+/\d+)\s*Break point\s*(\d+/\d+)', testo, re.IGNORECASE)
    if break_match:
        break_point = [break_match.group(1), break_match.group(2)]
    else:
        break_point = ["0/0", "0/0"]

    # break_point = [br1, br2]
    return ace, doppi_falli, break_point

def processa_match(testo_match, lista_tennisti):
    pattern_nomi = r'\b\w+\b'
    nomi_trovati = re.findall(pattern_nomi, testo_match)
    nomi_candidati = [n.lower() for n in nomi_trovati if len(n) > 3]

    trovati = trova_cognome_nella_lista(player_surname, nomi_candidati)

    if len(trovati) == 0:
        return None

    if len(trovati) == 1:
        giocatori = [trovati[0], "NON_RICONOSCIUTO"]
    else:
        giocatori = trovati[:2]

    # game_g1, game_g2 = estrai_game_da_testo(testo_match, giocatori)
    game_g1, game_g2 = estrai_game_da_testo(testo_match)
    ace, doppi_falli, break_point = estrai_statistiche(testo_match)

    risultati = []

    for idx, giocatore in enumerate(giocatori):
        break_player = break_point[idx] if idx < len(break_point) else "0/0"
        game_player = game_g1 if idx == 0 else game_g2
        game_avv = game_g2 if idx == 0 else game_g1

        risultati.append({
            "Giocatore": giocatore,
            "TOT GAME": sum(game_g1) + sum(game_g2),
            "TOT GAME PLAYER": sum(game_player),
            "DF": doppi_falli[idx],
            "BREAK": break_player,
            "ACE": ace[idx],
            "HND": sum(game_player) - sum(game_avv),
            "TIE BREAK": calcola_tie_break(game_g1, game_g2),
            "TORNEO": ""
        })

    return pd.DataFrame(risultati)

async def scrittura_in_excel(df_match, update):
    # Normalizza colonne
    df_match = df_match.rename(columns={
        "Giocatore": "GIOCATORE"
    })
    df_match = df_match[df_match["GIOCATORE"] != "NON_RICONOSCIUTO"]

    colonne_finali = [
        "GIOCATORE",
        "TOT GAME",
        "TOT GAME PLAYER",
        "DF",
        "BREAK",
        "ACE",
        "HND",
        "TIE BREAK",
        "TORNEO"
    ]
    df_match = df_match[colonne_finali]

    # Se file NON esiste → crealo con 2 sheet
    if not os.path.exists(EXCEL_LOCAL_PATH):
        await update.message.reply_text("📁 Creo nuovo Excel")

        with pd.ExcelWriter(EXCEL_LOCAL_PATH, engine="openpyxl") as writer:
            pd.DataFrame().to_excel(writer, sheet_name="Indice", index=False)
            df_match.to_excel(writer, sheet_name="Statistiche", index=False)

        return

    # File esiste → append su Statistiche
    df_esistente = pd.read_excel(EXCEL_LOCAL_PATH, sheet_name="Statistiche")

    df_aggiornato = pd.concat(
        [df_esistente, df_match],
        ignore_index=True
    )
    
    # ORDINA PER NOME GIOCATORE
    df_aggiornato = df_aggiornato.sort_values(
        by="GIOCATORE",
        ascending=True,
        kind="stable"
        )
    
    base, ext = os.path.splitext(EXCEL_LOCAL_PATH)
    temp_path = base + "_tmp" + ext
    
    with pd.ExcelWriter(
        temp_path,
        engine="openpyxl"
    ) as writer:
        pd.read_excel(EXCEL_LOCAL_PATH, sheet_name="Indice").to_excel(
            writer, sheet_name="Indice", index=False
        )
        df_aggiornato.to_excel(
            writer, sheet_name="Statistiche", index=False
        )
    
    # SOLO SE ARRIVA QUI → rimpiazza
    os.replace(temp_path, EXCEL_LOCAL_PATH)
    applica_schema_excel(
        EXCEL_LOCAL_PATH,
        "Statistiche",
        EXCEL_SCHEMA
    )


# ============================================================================
# GOOGLE DRIVE FUNCTIONS
# ============================================================================

# Download Excel da Drive
def download_excel_from_drive():
    request = drive_service.files().get_media(fileId=EXCEL_FILE_ID)
    fh = io.FileIO(EXCEL_LOCAL_PATH, "wb")
    downloader = MediaIoBaseDownload(fh, request)

    done = False
    while not done:
        _, done = downloader.next_chunk()

    print("✅ Excel scaricato da Drive")

# Upload excel in Drive
def upload_excel_to_drive():
    media = MediaFileUpload(
        EXCEL_LOCAL_PATH,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True
    )

    drive_service.files().update(
        fileId=EXCEL_FILE_ID,
        media_body=media
    ).execute()

    print("✅ Excel caricato su Drive")

EXCEL_SCHEMA = {
    "GIOCATORE": "text",
    "TOT GAME": "int",
    "TOT GAME PLAYER": "int",
    "DF": "int",
    "BREAK": "text",
    "ACE": "int",
    "HND": "int",
    "TIE BREAK": "int",
    "TORNEO": "text",
}

def applica_schema_excel(path_excel, sheet, schema):
    wb = load_workbook(path_excel)
    ws = wb[sheet]

    header = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}

    for col_name, col_type in schema.items():
        if col_name not in header:
            continue

        col_idx = header[col_name]

        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            cell = row[0]

            if col_type == "text":
                cell.number_format = "@"
                if cell.value is not None:
                    cell.value = str(cell.value)

            elif col_type == "int":
                cell.number_format = "0"
                try:
                    cell.value = int(cell.value)
                except:
                    cell.value = 0

    wb.save(path_excel)
    wb.close()

# ============================================================================
# TELEGRAM BOT HANDLERS
# ============================================================================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Messaggio di benvenuto"""
    await update.message.reply_text(
        "🎾 *Benvenuto nel Tennis Match Processor!*\n\n"
        "📸 Inviami una foto del tabellino del match e aggiornerò automaticamente il database.\n\n"
        "Comandi disponibili:\n"
        "/start - Mostra questo messaggio\n"
        "/help - Aiuto",
        parse_mode='Markdown'
    )

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Messaggio di aiuto"""
    await update.message.reply_text(
        "❓ *Come usare il bot:*\n\n"
        "1️⃣ Scatta una foto del tabellino del match\n"
        "2️⃣ Inviamela qui\n"
        "3️⃣ Aspetta qualche secondo\n"
        "4️⃣ Riceverai conferma dell'aggiornamento!\n\n"
        "✅ Il database sarà aggiornato automaticamente su Google Drive.",
        parse_mode='Markdown'
    )

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Gestisce le foto ricevute"""

    user_id = update.effective_user.id
    if user_id not in ALLOWED_USERS:
        await update.message.reply_text("⛔ Non sei autorizzato a usare questo bot.")
        return

    try:
        # Messaggio di attesa
        await update.message.reply_text("📸 Foto ricevuta! Estrazione testo...")
        
        # Download foto
        photo_file = await update.message.photo[-1].get_file()
        photo_path = "temp_match.jpg"
        await photo_file.download_to_drive(photo_path)
        print("DEBUG: immagine salvata in:", photo_path)
        
        from PIL import Image
        img = Image.open(photo_path)
        print("DEBUG: formato immagine:", img.format, img.mode, img.size)

        # Preprocessing immagine
        img = Image.open(photo_path)
        img = gray_scale_img(img)
        
        # OCR
        print("DEBUG: sto per chiamare pytesseract.image_to_string")
        try:
            text = pytesseract.image_to_string(
                img,
                lang="ita+eng",
                config="--oem 3 --psm 6"  #"--psm 6"
            )
            text = separa_maiuscole(text)
            print("DEBUG: OCR completato")
        except Exception as e:
            print("❌ ERRORE DURANTE OCR")
            traceback.print_exc()
            raise
        
        # Processa match
        await update.message.reply_text("⚙️ Processamento dati...")
        df_match = processa_match(text, player_surname)

        await update.message.reply_text(f"testo: {text}")
        if df_match is None or len(df_match) == 0:
            await update.message.reply_text("❌ Non sono riuscito a identificare i giocatori. Riprova con un'immagine più chiara.")
            return
        giocatori_trovati = df_match["Giocatore"].tolist()

        if len(giocatori_trovati) == 1:
            g1 = giocatori_trovati[0]
            stats1 = df_match.iloc[0]
        
            await update.message.reply_text(
                f"🎾 *{g1.upper()}*\n"
                f"   • Tot Game: {stats1['TOT GAME']}\n"
                f"   • Game: {stats1['TOT GAME PLAYER']}\n"
                f"   • DF: {stats1['DF']}\n"
                f"   • Break: {stats1['BREAK']}\n"
                f"   • Ace: {stats1['ACE']}\n"
                f"   • Handicap: {stats1['HND']:+d}\n\n"
                f"   • Tie Break: {stats1['TIE BREAK']}\n"
                f"⚠️ Altro giocatore NON RICONOSCIUTO",
                parse_mode="Markdown"
            )
        else:
            g1, g2 = giocatori_trovati
            stats1 = df_match[df_match['Giocatore'] == g1].iloc[0]
            stats2 = df_match[df_match['Giocatore'] == g2].iloc[0]
        
            await update.message.reply_text(
                f"🎾 *{g1.upper()}*\n"
                f"   • Tot Game: {stats1['TOT GAME']}\n"
                f"   • Game: {stats1['TOT GAME PLAYER']}\n"
                f"   • DF: {stats1['DF']}\n"
                f"   • Break: {stats1['BREAK']}\n"
                f"   • Ace: {stats1['ACE']}\n"
                f"   • Handicap: {stats1['HND']:+d}\n\n"
                f"   • Tie Break: {stats1['TIE BREAK']}\n"
                f"🎾 *{g2.upper()}*\n"
                f"   • Tot Game: {stats2['TOT GAME']}\n"
                f"   • Game: {stats2['TOT GAME PLAYER']}\n"
                f"   • DF: {stats2['DF']}\n"
                f"   • Break: {stats2['BREAK']}\n"
                f"   • Ace: {stats2['ACE']}\n"
                f"   • Handicap: {stats2['HND']:+d}\n\n"
                f"   • Tie Break: {stats2['TIE BREAK']}\n",
                parse_mode="Markdown"
            )
        
        # Salva in Excel
        download_excel_from_drive()
        
        await scrittura_in_excel(df_match, update)
        
        upload_excel_to_drive()
        
        # Messaggio di conferma
        # print("DEBUG: giocatori estratti =", giocatori)
        # print("DEBUG: len(giocatori) =", len(giocatori))
        # if len(giocatori) != 2:
        #     await update.message.reply_text(
        #         f"❌ Numero giocatori non valido ({len(giocatori)}): {giocatori}"
        #     )
        #     return
        
        # giocatori_trovati = df_match["Giocatore"].tolist()
        
        # if len(giocatori_trovati) == 2:
        #     g1, g2 = giocatori_trovati
        #     stats1 = df_match[df_match['Giocatore'] == g1].iloc[0]
        #     stats2 = df_match[df_match['Giocatore'] == g2].iloc[0]
        
        #     await update.message.reply_text(
        #         f"✅ *Match salvato con successo!*\n\n"
        #         f"🎾 *{g1.upper()}*\n"
        #         f"   • Game: {stats1['TOT GAME PLAYER']}\n"
        #         f"   • Ace: {stats1['ACE']}\n"
        #         f"   • DF: {stats1['DF']}\n"
        #         f"   • Handicap: {stats1['HND']:+d}\n\n"
        #         f"🎾 *{g2.upper()}*\n"
        #         f"   • Game: {stats2['TOT GAME PLAYER']}\n"
        #         f"   • Ace: {stats2['ACE']}\n"
        #         f"   • DF: {stats2['DF']}\n"
        #         f"   • Handicap: {stats2['HND']:+d}\n\n"
        #         f"💾 Database aggiornato!",
        #         parse_mode="Markdown"
        #     )
        
        # Cleanup
        os.remove(photo_path)
        estrai_game_da_testo
    except Exception as e:
        await update.message.reply_text(
            f"❌ Errore durante il processamento:\n{type(e).__name__}: {e}"
        )
        raise


async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Gestisce messaggi di testo"""
    await update.message.reply_text(
        "📸 Inviami una *foto* del tabellino, non testo!\n\n"
        "Usa /help per maggiori informazioni.",
        parse_mode='Markdown'
    )

# ============================================================================
# MAIN
# ============================================================================

def main():
    application = Application.builder().token(TELEGRAM_TOKEN).build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    PORT = int(os.environ.get("PORT", 10000))
    RENDER_URL = "https://tennis-fragolino-bot-1.onrender.com"

    print("🤖 Bot avviato in modalità WEBHOOK")

    application.run_webhook(
        listen="0.0.0.0",
        port=PORT,
        url_path="/",
        webhook_url=RENDER_URL
    )


if __name__ == '__main__':
    main()
































